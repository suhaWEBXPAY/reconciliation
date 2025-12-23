import re
import json
from datetime import datetime, date
import pandas as pd
import numpy as np
from typing import Tuple, Dict, List, Optional

from openpyxl.styles import PatternFill

"""Reconciliation engine

This file originally assumed fixed column names (hardcoded).

It now supports *prompt-driven* (AI) column mapping via the `mapping` argument
passed into `prep_portal` / `prep_bank`.

If `mapping` is not provided, it falls back to the old expected column names.
"""

# ---- LEGACY DEFAULT COLUMN NAMES (fallback) ----
DEFAULT_PORTAL_COLS = {
  "MID": "MID",
  "TID": "TID",
  "CARD NUMBER": "CARD NUMBER",
  "AMOUNT": "AMOUNT",
}

DEFAULT_BANK_COLS = {
  "MERCHANTID": "MERCHANTID",
  "TID": "TID",
  "CARDNUMBER_MASKED": "CARDNUMBER_MASKED",
  "GROSSSALES": "GROSSSALES",
}

def _canon(s: str) -> str:
  return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def _find_col(df: pd.DataFrame, wanted: str) -> str:
  # exact match first, else canonical match
  if wanted in df.columns:
    return wanted
  cmap = { _canon(c): c for c in df.columns }
  key = _canon(wanted)
  return cmap.get(key, "")

def _find_col_any(df: pd.DataFrame, candidates: List[str]) -> str:
  """Try multiple candidate column names and return the first match."""
  for c in candidates:
    found = _find_col(df, c)
    if found:
      return found
  return ""

def _norm_id(v) -> str:
  if pd.isna(v):
    return ""
  if isinstance(v, (int, np.integer)):
    return str(int(v))
  if isinstance(v, (float, np.floating)) and abs(v - round(v)) < 1e-9:
    return str(int(round(v)))
  s = str(v).strip()
  # strip trailing .0 etc
  if re.fullmatch(r"\d+\.0+", s):
    s = s.split(".")[0]
  return s

def _card_first_last(card) -> Tuple[str, str]:
  if pd.isna(card):
    return ("", "")
  digits = re.findall(r"\d", str(card))
  if not digits:
    return ("", "")
  first4 = "".join(digits[:4]) if len(digits) >= 4 else "".join(digits)
  last4 = "".join(digits[-4:]) if len(digits) >= 4 else "".join(digits)
  return (first4, last4)

def _norm_amount(a):
  if pd.isna(a):
    return None
  try:
    return float(round(float(a) + 1e-9, 2))
  except Exception:
    return None

def read_excel_any(path_or_file) -> pd.DataFrame:
  # More reliable than plain pd.read_excel() (prevents "engine cannot be determined" in many cases)
  try:
    return pd.read_excel(path_or_file, engine="openpyxl")
  except Exception:
    # fallback for environments that still infer engine successfully
    return pd.read_excel(path_or_file)

def _resolve_required_cols(
  df: pd.DataFrame,
  required: Dict[str, str],
  mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
  """Resolve logical fields -> actual dataframe column names.

  required: logical -> legacy default column label (used for fallback lookup)
  mapping: logical -> column name provided by AI / prompt
  """
  mapping = mapping or {}

  cols: Dict[str, str] = {}
  missing: List[str] = []
  for logical, legacy_label in required.items():
    # 1) AI mapping (exact)
    ai_col = (mapping.get(logical) or "").strip()
    if ai_col and ai_col in df.columns:
      cols[logical] = ai_col
      continue

    # 2) AI mapping (canonical match)
    if ai_col:
      found = _find_col(df, ai_col)
      if found:
        cols[logical] = found
        continue

    # 3) Legacy default label fallback
    found = _find_col(df, legacy_label)
    if found:
      cols[logical] = found
      continue

    missing.append(f"{logical} (expected like '{legacy_label}')")

  if missing:
    raise ValueError(
      "Missing required fields: "
      + ", ".join(missing)
      + f". Found columns: {list(df.columns)}"
    )
  return cols


def prep_portal(df: pd.DataFrame, source_file: str, mapping: Optional[Dict[str, str]] = None) -> pd.DataFrame:
  """Prepare portal dataframe.

  mapping format (logical -> column name):
    {"mid": "...", "tid": "...", "card": "...", "amount": "..."}
  """
  df = df.copy()

  # ✅ FILTER RULES:
  # 1) Ignore any TRNX Type containing "amex" (covers "amex" and "void amex")
  # 2) For "void sale": ignore the void row itself AND ignore the corresponding SALE row
  #    by matching RRN + AUTH (so we don't accidentally drop a later legitimate sale with same mid/tid/card/amount).
  trnx_col = _find_col(df, "TRNX Type")
  if trnx_col:
    s = df[trnx_col].astype(str).str.strip()

    # (A) Drop any AMEX-related rows outright
    amex_mask = s.str.contains(r"\bamex\b", case=False, na=False)
    if amex_mask.any():
      df = df[~amex_mask].copy()
      # refresh s after filtering
      s = df[trnx_col].astype(str).str.strip()

    # (B) Handle VOID SALE pairing using RRN + AUTH
    void_mask = s.str.contains(r"\bvoid\b", case=False, na=False)

    if void_mask.any():
      # Try to locate RRN and AUTH-like columns (portal naming varies)
      rrn_col = _find_col_any(df, [
        "RRN",
        "Retrieval Reference Number",
        "Retrieval Ref No",
        "Retrieval Ref",
        "RRN No",
      ])
      auth_col = _find_col_any(df, [
        "AUTH",
        "Auth",
        "AUTH CODE",
        "Auth Code",
        "Authorization Code",
        "Approval Code",
        "Auth No",
      ])

      if rrn_col and auth_col:
        rrn_norm = df[rrn_col].apply(_norm_id)
        auth_norm = df[auth_col].apply(_norm_id)

        # Collect (RRN, AUTH) pairs from VOID rows
        void_pairs = set(
          (r, a)
          for r, a in zip(rrn_norm[void_mask], auth_norm[void_mask])
          if r and a
        )

        # Drop the VOID rows themselves
        keep = ~void_mask

        # Also drop non-void rows whose (RRN, AUTH) matches a void pair
        non_void = ~void_mask
        match_void_pair = non_void & rrn_norm.astype(str).ne("") & auth_norm.astype(str).ne("") & list(
          (r, a) in void_pairs for r, a in zip(rrn_norm, auth_norm)
        )

        keep = keep & ~match_void_pair
        df = df[keep].copy()

      else:
        # If we can't find RRN/AUTH columns, safest fallback:
        # drop VOID rows only (do NOT attempt to drop a "matching sale" by mid/tid/card/amount)
        df = df[~void_mask].copy()

  cols = _resolve_required_cols(
    df,
    required={
      "mid": DEFAULT_PORTAL_COLS["MID"],
      "tid": DEFAULT_PORTAL_COLS["TID"],
      "card": DEFAULT_PORTAL_COLS["CARD NUMBER"],
      "amount": DEFAULT_PORTAL_COLS["AMOUNT"],
    },
    mapping=mapping,
  )

  out = pd.DataFrame({
    "mid": df[cols["mid"]],
    "tid": df[cols["tid"]],
    "card_raw": df[cols["card"]],
    "amount": df[cols["amount"]],
  })

  # Store the full original row for dynamic export (no hardcoded columns)
  raw_cols = list(df.columns)
  out["raw_cols"] = json.dumps(raw_cols)
  # Build JSON per row (convert numpy types to python primitives)
  out["raw_json"] = df.apply(
    lambda r: json.dumps(
      {k: (None if pd.isna(v) else (v.item() if hasattr(v, "item") else v)) for k, v in r.to_dict().items()},
      default=str
    ),
    axis=1
  )

  out["mid_norm"] = out["mid"].apply(_norm_id)
  out["tid_norm"] = out["tid"].apply(_norm_id)
  fl = out["card_raw"].apply(_card_first_last)
  out["card_first4"] = fl.apply(lambda x: x[0])
  out["card_last4"] = fl.apply(lambda x: x[1])
  out["amount_norm"] = out["amount"].apply(_norm_amount)

  out["match_key"] = out.apply(
    lambda r: f'{r["mid_norm"]}|{r["tid_norm"]}|{r["card_first4"]}|{r["card_last4"]}|{r["amount_norm"]}',
    axis=1,
  )
  out["source_file"] = source_file
  return out[["match_key","mid_norm","tid_norm","card_raw","card_first4","card_last4","amount_norm","source_file","raw_cols","raw_json"]]

def prep_bank(df: pd.DataFrame, source_file: str, mapping: Optional[Dict[str, str]] = None) -> pd.DataFrame:
  """Prepare bank dataframe.

  mapping format (logical -> column name):
    {"mid": "...", "tid": "...", "card": "...", "amount": "..."}
  """
  df = df.copy()

  # ✅ BANK FILTER RULES (YOUR REQUIREMENT):
  # 1) Ignore completely empty rows anywhere in the sheet (including empty rows in between)
  # 2) Ignore rows where ONLY "WEBXPAY INCOME" has a value (everything else blank)

  # Treat whitespace-only cells as empty, then drop fully empty rows
  df = df.replace(r"^\s*$", pd.NA, regex=True)
  df = df.dropna(how="all").copy()

  # Find "WEBXPAY INCOME" column (bank files may vary slightly in header text)
  income_col = _find_col_any(df, [
    "WEBXPAY INCOME",
    "WEBXPAY INCOME (LKR)",
    "WEBXPAY INCOME LKR",
    "WEBXPAY Income",
    "Webxpay Income",
    "WEBXPAY_INCOME",
  ])

  if income_col:
    other_cols = [c for c in df.columns if c != income_col]
    if other_cols:
      # Row is junk if income is present AND every other column is empty
      mask_only_income = df[income_col].notna() & df[other_cols].isna().all(axis=1)
      if mask_only_income.any():
        df = df.loc[~mask_only_income].copy()

  cols = _resolve_required_cols(
    df,
    required={
      "mid": DEFAULT_BANK_COLS["MERCHANTID"],
      "tid": DEFAULT_BANK_COLS["TID"],
      "card": DEFAULT_BANK_COLS["CARDNUMBER_MASKED"],
      "amount": DEFAULT_BANK_COLS["GROSSSALES"],
    },
    mapping=mapping,
  )

  out = pd.DataFrame({
    "merchantid": df[cols["mid"]],
    "tid": df[cols["tid"]],
    "card_masked": df[cols["card"]],
    "amount": df[cols["amount"]],
  })

  # Store the full original row for dynamic export (no hardcoded columns)
  raw_cols = list(df.columns)
  out["raw_cols"] = json.dumps(raw_cols)
  out["raw_json"] = df.apply(
    lambda r: json.dumps(
      {k: (None if pd.isna(v) else (v.item() if hasattr(v, "item") else v)) for k, v in r.to_dict().items()},
      default=str
    ),
    axis=1
  )

  out["mid_norm"] = out["merchantid"].apply(_norm_id)
  out["tid_norm"] = out["tid"].apply(_norm_id)
  fl = out["card_masked"].apply(_card_first_last)
  out["card_first4"] = fl.apply(lambda x: x[0])
  out["card_last4"] = fl.apply(lambda x: x[1])
  out["amount_norm"] = out["amount"].apply(_norm_amount)

  out["match_key"] = out.apply(
    lambda r: f'{r["mid_norm"]}|{r["tid_norm"]}|{r["card_first4"]}|{r["card_last4"]}|{r["amount_norm"]}',
    axis=1,
  )
  out["source_file"] = source_file
  return out[["match_key","mid_norm","tid_norm","card_masked","card_first4","card_last4","amount_norm","source_file","raw_cols","raw_json"]]

def _pool_from_rows(rows: List[dict]) -> Dict[str, List[int]]:
  pool: Dict[str, List[int]] = {}
  for r in rows:
    pool.setdefault(r["match_key"], []).append(r["id"])
  return pool

def reconcile_and_update_db(conn, new_portal: pd.DataFrame, new_bank: pd.DataFrame, user_id: int, run_id: Optional[int] = None):
  """Consumes new rows and updates open_* tables in the DB.
  - Deletes matched old rows
  - Inserts unmatched new rows
  Returns stats dict.
  """
  # Load existing open rows from DB
  open_portal_rows = conn.execute(
    "SELECT id, match_key FROM open_portal WHERE user_id=?",
    (int(user_id),)
  ).fetchall()
  open_bank_rows = conn.execute(
    "SELECT id, match_key FROM open_bank WHERE user_id=?",
    (int(user_id),)
  ).fetchall()
  open_portal_pool = _pool_from_rows([dict(r) for r in open_portal_rows])
  open_bank_pool   = _pool_from_rows([dict(r) for r in open_bank_rows])

  # Build pool for new bank rows (not in DB yet)
  new_bank_pool: Dict[str, List[int]] = {}
  for idx, r in new_bank.reset_index(drop=True).iterrows():
    new_bank_pool.setdefault(r["match_key"], []).append(idx)

  matched_this_run = []

  # 1) Match new portal vs new bank first (same upload)
  portal_unmatched_idx = []
  for pidx, prow in new_portal.reset_index(drop=True).iterrows():
    key = prow["match_key"]
    if key in new_bank_pool and new_bank_pool[key]:
      bidx = new_bank_pool[key].pop()
      matched_this_run.append(("new_portal", pidx, "new_bank", bidx, key))
    else:
      portal_unmatched_idx.append(pidx)

  # Build remaining new bank unmatched idx
  bank_unmatched_idx = []
  for key, lst in new_bank_pool.items():
    bank_unmatched_idx.extend(lst)

  # 2) Match remaining new portal vs OPEN BANK (from previous days)
  final_portal_unmatched = []
  bank_ids_to_delete = []
  for pidx in portal_unmatched_idx:
    key = new_portal.iloc[pidx]["match_key"]
    if key in open_bank_pool and open_bank_pool[key]:
      bid = open_bank_pool[key].pop()
      bank_ids_to_delete.append(bid)
      matched_this_run.append(("new_portal", pidx, "open_bank", bid, key))
    else:
      final_portal_unmatched.append(pidx)

  # 3) Match remaining new bank vs OPEN PORTAL (from previous days)
  final_bank_unmatched = []
  portal_ids_to_delete = []
  for bidx in bank_unmatched_idx:
    key = new_bank.iloc[bidx]["match_key"]
    if key in open_portal_pool and open_portal_pool[key]:
      pid = open_portal_pool[key].pop()
      portal_ids_to_delete.append(pid)
      matched_this_run.append(("new_bank", bidx, "open_portal", pid, key))
    else:
      final_bank_unmatched.append(bidx)

  # Apply deletions (and log rows for undo)
  def _log_deleted_portal(ids: List[int]):
    if not run_id or not ids:
      return
    qs = ",".join(["?"] * len(ids))
    rows = conn.execute(
      f"SELECT id, match_key, mid, tid, card_raw, card_first4, card_last4, amount, source_file, raw_cols, raw_json, inserted_at FROM open_portal WHERE user_id=? AND id IN ({qs})",
      [int(user_id)] + [int(i) for i in ids]
    ).fetchall()
    if not rows:
      return
    conn.executemany(
      """INSERT OR REPLACE INTO recon_deleted_portal(
           run_id, user_id, orig_id, match_key, mid, tid, card_raw, card_first4, card_last4,
           amount, source_file, raw_cols, raw_json, inserted_at
         ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      [
        (
          int(run_id), int(user_id), int(r["id"]), r["match_key"], r["mid"], r["tid"], r["card_raw"],
          r["card_first4"], r["card_last4"], r["amount"], r["source_file"], r["raw_cols"], r["raw_json"], r["inserted_at"]
        )
        for r in rows
      ]
    )

  def _log_deleted_bank(ids: List[int]):
    if not run_id or not ids:
      return
    qs = ",".join(["?"] * len(ids))
    rows = conn.execute(
      f"SELECT id, match_key, merchantid, tid, card_masked, card_first4, card_last4, amount, source_file, raw_cols, raw_json, inserted_at FROM open_bank WHERE user_id=? AND id IN ({qs})",
      [int(user_id)] + [int(i) for i in ids]
    ).fetchall()
    if not rows:
      return
    conn.executemany(
      """INSERT OR REPLACE INTO recon_deleted_bank(
           run_id, user_id, orig_id, match_key, merchantid, tid, card_masked, card_first4, card_last4,
           amount, source_file, raw_cols, raw_json, inserted_at
         ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      [
        (
          int(run_id), int(user_id), int(r["id"]), r["match_key"], r["merchantid"], r["tid"], r["card_masked"],
          r["card_first4"], r["card_last4"], r["amount"], r["source_file"], r["raw_cols"], r["raw_json"], r["inserted_at"]
        )
        for r in rows
      ]
    )

  _log_deleted_portal(portal_ids_to_delete)
  _log_deleted_bank(bank_ids_to_delete)

  if portal_ids_to_delete:
    conn.executemany(
      "DELETE FROM open_portal WHERE user_id=? AND id=?",
      [(int(user_id), i) for i in portal_ids_to_delete]
    )
  if bank_ids_to_delete:
    conn.executemany(
      "DELETE FROM open_bank WHERE user_id=? AND id=?",
      [(int(user_id), i) for i in bank_ids_to_delete]
    )

  # Insert remaining unmatched new portal rows into DB
  if final_portal_unmatched:
    rows = []
    for pidx in final_portal_unmatched:
      r = new_portal.iloc[pidx]
      rows.append((
        int(user_id),
        int(run_id) if run_id else None,
        r["match_key"],
        r["mid_norm"],
        r["tid_norm"],
        r["card_raw"],
        r["card_first4"],
        r["card_last4"],
        r["amount_norm"],
        r["source_file"],
        r.get("raw_cols"),
        r.get("raw_json"),
      ))
    conn.executemany(
      """INSERT INTO open_portal(
            user_id, run_id, match_key, mid, tid, card_raw, card_first4, card_last4, amount, source_file,
            raw_cols, raw_json
          ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
      rows
    )

  # Insert remaining unmatched new bank rows into DB
  if final_bank_unmatched:
    rows = []
    for bidx in final_bank_unmatched:
      r = new_bank.iloc[bidx]
      rows.append((
        int(user_id),
        int(run_id) if run_id else None,
        r["match_key"],
        r["mid_norm"],
        r["tid_norm"],
        r["card_masked"],
        r["card_first4"],
        r["card_last4"],
        r["amount_norm"],
        r["source_file"],
        r.get("raw_cols"),
        r.get("raw_json"),
      ))
    conn.executemany(
      """INSERT INTO open_bank(
            user_id, run_id, match_key, merchantid, tid, card_masked, card_first4, card_last4, amount, source_file,
            raw_cols, raw_json
          ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
      rows
    )

  return {
    "matched_count": len(matched_this_run),
    "new_portal_rows": int(len(new_portal)),
    "new_bank_rows": int(len(new_bank)),
    "inserted_portal": int(len(final_portal_unmatched)),
    "inserted_bank": int(len(final_bank_unmatched)),
    "deleted_open_portal": int(len(portal_ids_to_delete)),
    "deleted_open_bank": int(len(bank_ids_to_delete)),
  }

def export_unreconciled(conn, out_path: str, user_id: int):
  """Export unreconciled rows.

  Requirements:
  - Export columns must NOT be hardcoded; use the exact uploaded columns.
  - Add a column showing the date it first became unreconciled.
  - Highlight rows: >=7 days (yellow), >=14 days (red).
  """
  from openpyxl.styles import PatternFill

  def _safe_date_only(inserted_at: str) -> str:
    # inserted_at is typically 'YYYY-MM-DD HH:MM:SS'
    if not inserted_at:
      return ""
    return str(inserted_at)[:10]

  def _build_df(table: str) -> pd.DataFrame:
    rows = conn.execute(
      f"SELECT raw_cols, raw_json, source_file, inserted_at FROM {table} WHERE user_id=? ORDER BY inserted_at DESC",
      (int(user_id),)
    ).fetchall()
    if not rows:
      return pd.DataFrame()

    # Determine column order (preserve first-seen order across files)
    col_order: List[str] = []
    seen = set()
    parsed_rows = []
    for rr in rows:
      raw_cols = []
      if rr["raw_cols"]:
        try:
          raw_cols = json.loads(rr["raw_cols"])
        except Exception:
          raw_cols = []
      for c in raw_cols:
        if c not in seen:
          seen.add(c)
          col_order.append(c)

      payload = {}
      if rr["raw_json"]:
        try:
          payload = json.loads(rr["raw_json"])
        except Exception:
          payload = {}
      payload = dict(payload)
      payload["SOURCE_FILE"] = rr["source_file"]
      payload["FIRST_UNRECONCILED_DATE"] = _safe_date_only(rr["inserted_at"])
      parsed_rows.append(payload)

    # Ensure metadata columns at the end
    for c in ["SOURCE_FILE", "FIRST_UNRECONCILED_DATE"]:
      if c not in seen:
        col_order.append(c)

    df = pd.DataFrame(parsed_rows)
    # Reindex to enforce order (missing columns become NaN)
    return df.reindex(columns=col_order)

  portal_df = _build_df("open_portal")
  bank_df   = _build_df("open_bank")

  yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
  red    = PatternFill(start_color="FF8A80", end_color="FF8A80", fill_type="solid")

  def _apply_aging(ws, first_date_col_idx: int):
    today = date.today()
    # header is row 1
    for r in range(2, ws.max_row + 1):
      v = ws.cell(row=r, column=first_date_col_idx).value
      if not v:
        continue
      try:
        d0 = datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
      except Exception:
        continue
      age = (today - d0).days
      fill = red if age >= 14 else (yellow if age >= 7 else None)
      if fill is None:
        continue
      for c in range(1, ws.max_column + 1):
        ws.cell(row=r, column=c).fill = fill

  with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    portal_df.to_excel(writer, index=False, sheet_name="Unreconciled_Portal")
    bank_df.to_excel(writer, index=False, sheet_name="Unreconciled_Bank")

    wb = writer.book
    for sheet_name, df in [("Unreconciled_Portal", portal_df), ("Unreconciled_Bank", bank_df)]:
      ws = wb[sheet_name]
      if df.empty:
        continue
      # find FIRST_UNRECONCILED_DATE column index (1-based)
      try:
        col_idx = list(df.columns).index("FIRST_UNRECONCILED_DATE") + 1
      except ValueError:
        continue
      _apply_aging(ws, col_idx)
